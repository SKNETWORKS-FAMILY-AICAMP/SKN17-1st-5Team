from openpyxl import load_workbook
from db_connector import get_connection
import os

def registed_car():
    connection = get_connection()
    cursor = connection.cursor()

    car_type = ['승용', '승합', '화물', '특수']

    # 엑셀 파일 열기 만약 여러 줄이면 여기서부터 반복
    folder_path = './/data'  # 원하는 폴더 경로
    file_list = os.listdir(folder_path)
    for file_name in file_list: # 파일 이름 하나씩 읽기
        if ".xlsx" not in file_name: continue # xlsx 파일 아니면 넘기기
        wb = load_workbook(folder_path+"/" + file_name)

        sheet = wb.worksheets[1]  # n-0 번째 시트
        city = ''
        relationship = {}
        # 셀 하나씩
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if idx == 2: year, mon = row[1].split('.')
            if idx < 5: continue
            if row[0] == '합계': break
            town = row[1]
            if town == "계": continue
            # if " " in town : town = town.split()[0]
            if row[0] != None : city = row[0]
            for t, c in enumerate([row[5],row[9],row[13],row[17]]):
                key = city + '_' + town + '_' + car_type[t]
                try:
                    relationship[key] += c
                except:
                    relationship[key] = c

        for key in relationship.keys():
            query_value = str(year) + ", " + str(mon) + ", '" +  key.replace("_", "', '") + "', " + str(relationship[key])
            query = f"insert registed_car (added_year, added_month, city, town, type, total, local_regist) values ({query_value}, 'Y')"
            cursor.execute(query)


    connection.commit()
    cursor.close()
    connection.close()

registed_car()